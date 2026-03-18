"""Read or create Excel XLSX files."""

import json
import os


def run(path: str, operation: str, data: str = "", sheet: str = "", headers: str = "") -> str:
    """Read data from or create an Excel XLSX file.

    @param path: Path to the Excel file.
    @param operation: 'read' to extract data, 'create' to write a new file.
    @param data: JSON string of array of arrays (required for create).
    @param sheet: Sheet name to read (defaults to active sheet).
    @param headers: Comma-separated header names (optional, create only).
    @returns JSON data or confirmation message.
    @throws ValueError: If the operation is invalid.
    """
    if operation == "read":
        return _read_xlsx(path, sheet)
    if operation == "create":
        return _create_xlsx(path, data, headers)
    raise ValueError(f"Invalid operation '{operation}'. Must be 'read' or 'create'.")


def _read_xlsx(file_path: str, sheet: str) -> str:
    """Read all data from an Excel file and return it as JSON."""
    try:
        import openpyxl
    except ImportError:
        return ("error: openpyxl is required but not installed. "
                "Install it with: pip install openpyxl")

    resolved = os.path.abspath(file_path)
    if not os.path.exists(resolved):
        raise FileNotFoundError(f"Excel file not found: {resolved}")

    workbook = openpyxl.load_workbook(resolved, read_only=True, data_only=True)

    if sheet:
        if sheet not in workbook.sheetnames:
            available = ", ".join(workbook.sheetnames)
            workbook.close()
            raise ValueError(f"Sheet '{sheet}' not found. Available sheets: {available}")
        worksheet = workbook[sheet]
    else:
        worksheet = workbook.active

    rows: list[list[object]] = []
    for row in worksheet.iter_rows(values_only=True):
        rows.append([_serialize_cell(cell) for cell in row])

    workbook.close()

    if not rows:
        return "No data found in the spreadsheet."

    return json.dumps(rows, indent=2, default=str)


def _serialize_cell(value: object) -> object:
    """Convert a cell value to a JSON-safe type."""
    if value is None:
        return None
    if isinstance(value, (int, float, bool, str)):
        return value
    return str(value)


def _create_xlsx(file_path: str, data: str, headers: str) -> str:
    """Create an Excel file from JSON array of arrays."""
    try:
        import openpyxl
    except ImportError:
        return ("error: openpyxl is required but not installed. "
                "Install it with: pip install openpyxl")

    try:
        rows = json.loads(data)
    except json.JSONDecodeError as err:
        raise ValueError(f"Invalid JSON data: {err}")

    if not isinstance(rows, list):
        raise ValueError("Data must be a JSON array of arrays.")

    resolved = os.path.abspath(file_path)
    parent_dir = os.path.dirname(resolved)
    if parent_dir and not os.path.exists(parent_dir):
        os.makedirs(parent_dir, exist_ok=True)

    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    if headers:
        worksheet.append(headers.split(","))

    for row in rows:
        if not isinstance(row, list):
            raise ValueError(f"Each row must be an array, got {type(row).__name__}.")
        worksheet.append(row)

    workbook.save(resolved)
    return f"Excel file created successfully at: {resolved}"
